import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, numbers
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Weekly Inflight Report Generator", layout="wide")

st.title("📊 Weekly Inflight Report Generator")

# --- Configuration ---
ALL_BRANDS = [
    "VS", "BB", "HM", "CT", "AO", "RC", "MC", "PR", "SS", "UB", "FL",
    "CP", "AY", "MU", "AU", "BT", "CO", "NB", "PI", "HN", "TR", "CF"
]

ALL_MARKETS = ["UAE", "KSA", "KWT", "QAT", "EGY", "BAH", "JOR", "OMN"]

ALL_CHANNELS = [
    "Brand Paid Search", "Facebook", "GDN", "Non-Brand Paid Search",
    "Snapchat", "TikTok", "UAC", "YouTube"
]

ALL_FUNNELS = ["Awareness", "Consideration", "Conversion"]


# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

def find_header_row(sheet_df, keyword="campaign name"):
    """Find the row index that contains the header keywords."""
    for idx, row in sheet_df.iterrows():
        row_vals = [str(v).lower().strip() if pd.notna(v) else "" for v in row.values]
        if keyword.lower() in row_vals:
            return idx
    return 0


def normalize_campaign(name):
    """Normalize campaign name for matching."""
    if pd.isna(name):
        return ""
    return str(name).strip().lower()


def extract_placement(campaign_name):
    """Extract ad placement from campaign name (last segment after underscore)."""
    if pd.isna(campaign_name):
        return ""
    parts = str(campaign_name).strip().rsplit("_", 1)
    return parts[-1] if len(parts) > 1 else ""


def format_platform(channel, optimization, funnel, campaign_name="", combine=False):
    """Format platform name similar to the template.
    When combine=True, similar platforms are grouped together (e.g. all TikTok variants -> TIKTOK).
    """
    ch = str(channel).strip().upper()
    opt = str(optimization).strip().upper()
    camp = str(campaign_name).strip().lower()

    lang_suffix = ""
    if "english" in camp or "_en_" in camp or camp.endswith("-en"):
        lang_suffix = "-EN"
    elif "arabic" in camp or "_ar_" in camp or camp.endswith("-ar"):
        lang_suffix = "-AR"

    if ch in ["META", "FACEBOOK", "INSTAGRAM", "FB"]:
        return "IG" if combine else (f"IG{lang_suffix}" if funnel == "Consideration" and lang_suffix else "IG")
    elif ch in ["TIKTOK", "TIK TOK"]:
        if combine:
            return "TIKTOK"
        if "PULSE" in opt:
            return "TIKTOK PULSE"
        elif "SEARCH" in opt or "SEARCH ADS" in opt:
            return "TIKTOK SEARCH"
        elif "VIEW" in opt or "VC" in opt:
            return "TIKTOK VIEW"
        elif "REACH" in opt:
            return "TIKTOK REACH"
        return "TIKTOK"
    elif ch in ["SNAPCHAT", "SNAP"]:
        return "SNAPCHAT"
    elif ch in ["PINTEREST"]:
        return "PINTEREST"
    elif ch in ["YOUTUBE", "YT"]:
        return "YOUTUBE"
    elif ch in ["GDN"]:
        return "GDN"
    elif ch in ["BRAND PAID SEARCH", "PAID SEARCH"]:
        return "Paid Search" if combine else "Brand Paid Search"
    elif ch in ["NON-BRAND PAID SEARCH"]:
        return "Paid Search" if combine else "Non-Brand Paid Search"
    elif ch in ["UAC"]:
        return "UAC"
    return ch


def safe_numeric(val):
    """Safely convert a value to numeric, returning 0 if invalid."""
    try:
        v = pd.to_numeric(val, errors="coerce")
        return 0 if pd.isna(v) else float(v)
    except Exception:
        return 0


# =============================================================================
# FORMATTED REPORT BUILDER
# =============================================================================

def build_formatted_report(report_data_by_market, all_data):
    """Create a formatted workbook with the REPORT sheet matching the template style."""
    wb = Workbook()
    ws = wb.active
    ws.title = "REPORT"

    # --- Styles ---
    dark_red = "C00000"
    yellow = "FFFF00"
    gold = "FFC000"
    light_green = "C6EFCE"
    cyan = "00B0F0"
    gray = "D9D9D9"
    white = "FFFFFF"
    black = "000000"

    header_fill = PatternFill(start_color=dark_red, end_color=dark_red, fill_type="solid")
    header_font = Font(color=white, bold=True, size=9, name="Calibri")
    kpi_fill = PatternFill(start_color=yellow, end_color=yellow, fill_type="solid")
    kpi_font = Font(bold=True, size=10, color=black, name="Calibri")
    awareness_fill = PatternFill(start_color=gold, end_color=gold, fill_type="solid")
    awareness_font = Font(bold=True, size=9, color=black, name="Calibri")
    consideration_fill = PatternFill(start_color=light_green, end_color=light_green, fill_type="solid")
    consideration_font = Font(bold=True, size=9, color=black, name="Calibri")
    conversion_fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
    conversion_font = Font(bold=True, size=9, color=black, name="Calibri")
    platform_fill = PatternFill(start_color=cyan, end_color=cyan, fill_type="solid")
    platform_font = Font(color=white, bold=True, size=9, name="Calibri")
    market_fill = PatternFill(start_color=gray, end_color=gray, fill_type="solid")
    market_font = Font(bold=True, size=11, color=black, name="Calibri")
    data_font = Font(size=9, color=black, name="Calibri")
    thin_border = Border(
        left=Side(style="thin", color=black),
        right=Side(style="thin", color=black),
        top=Side(style="thin", color=black),
        bottom=Side(style="thin", color=black),
    )
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right_align = Alignment(horizontal="right", vertical="center")

    # --- Column widths ---
    col_widths = {
        1: 10,   # Market
        2: 12,   # Region
        3: 12,   # Audience
        4: 12,   # Ad Placement
        5: 12,   # Funnel
        6: 16,   # Platform
        7: 18,   # Campaign/Ad Format
        8: 12,   # Budget Planned
        9: 12,   # Budget Achieved
        10: 8,   # Budget %
        11: 14,  # Impressions Planned
        12: 14,  # Impressions Achieved
        13: 8,   # Impressions %
        14: 12,  # Reach/Clicks Planned
        15: 12,  # Reach/Clicks Achieved
        16: 8,   # Reach/Clicks %
        17: 10,  # CPM/CPC Planned
        18: 10,  # CPM/CPC Achieved
        19: 8,   # Freq/CTR Planned
        20: 8,   # Freq/CTR Achieved
        21: 10,  # Views/Sessions
        22: 10,  # Engagement/Purchases
        23: 8,   # VTR/Purchases Value
        24: 8,   # ER/Purchase Roas
        25: 10,  # Sessions (Awareness only)
        26: 10,  # Purchases
        27: 14,  # Purchases Value
        28: 12,  # Purchase Roas
    }
    for col, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    # --- Header definitions ---
    AWARENESS_HEADERS = [
        "Region", "Audience", "Ad Placement", "Funnel", "Platform", "Campaign",
        "Budget", "", "",
        "Impressions", "", "",
        "Reach", "", "",
        "CPM", "",
        "Freq", "",
        "Views", "Engagement", "VTR", "ER",
        "Sessions", "Purchases", "Purchases Value", "Purchase Roas",
    ]
    AWARENESS_SUBHEADERS = [
        "", "", "", "", "", "",
        "Planned", "Achieved", "%",
        "Planned", "Achieved", "%",
        "Planned", "Achieved", "%",
        "Planned", "Achieved",
        "Planned", "Achieved",
        "Achieved", "Achieved", "Achieved", "Achieved",
        "Achieved", "Achieved", "Achieved", "Achieved",
    ]

    CONSIDERATION_HEADERS = [
        "Region", "Audience", "Ad Placement", "Funnel", "Platform", "Ad Format",
        "Budget", "", "",
        "Impressions", "", "",
        "Clicks", "", "",
        "CPC", "",
        "CTR", "",
        "Sessions", "Purchases", "Purchases Value", "Purchase Roas",
    ]
    CONSIDERATION_SUBHEADERS = [
        "", "", "", "", "", "",
        "Planned", "Achieved", "%",
        "Planned", "Achieved", "%",
        "Planned", "Achieved", "%",
        "Planned", "Achieved",
        "Planned", "Achieved",
        "Achieved", "Achieved", "Achieved", "Achieved",
    ]

    def add_kpi_banner(row, last_col):
        cell = ws.cell(row=row, column=9, value="Primary KPI")
        cell.fill = kpi_fill
        cell.font = kpi_font
        cell.alignment = center_align
        cell.border = thin_border
        ws.merge_cells(start_row=row, start_column=9, end_row=row, end_column=last_col)
        for col in range(9, last_col + 1):
            c = ws.cell(row=row, column=col)
            c.fill = kpi_fill
            c.border = thin_border
        return row + 1

    def apply_header_row(row, headers, subheaders, last_col):
        # Main headers (start from col 2; col 1 is the merged market cell)
        for col, val in enumerate(headers, 2):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = center_align
        # Sub-headers
        for col, val in enumerate(subheaders, 2):
            cell = ws.cell(row=row + 1, column=col, value=val)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = center_align
        # Merge grouped headers
        ws.merge_cells(start_row=row, start_column=9, end_row=row, end_column=11)   # Budget
        ws.merge_cells(start_row=row, start_column=12, end_row=row, end_column=14)  # Impressions
        ws.merge_cells(start_row=row, start_column=15, end_row=row, end_column=17) # Reach/Clicks
        ws.merge_cells(start_row=row, start_column=18, end_row=row, end_column=19) # CPM/CPC
        ws.merge_cells(start_row=row, start_column=20, end_row=row, end_column=21) # Freq/CTR
        return row + 2

    def write_data_row(row, data, funnel_type):
        # Column 2: Region
        c2 = ws.cell(row=row, column=2, value=data.get("region", ""))
        c2.border = thin_border; c2.alignment = center_align; c2.font = data_font

        # Column 3: Audience
        c3 = ws.cell(row=row, column=3, value=data.get("audience", ""))
        c3.border = thin_border; c3.alignment = center_align; c3.font = data_font

        # Column 4: Ad Placement
        c4 = ws.cell(row=row, column=4, value=data.get("placement", ""))
        c4.border = thin_border; c4.alignment = center_align; c4.font = data_font

        # Column 5: Funnel
        funnel_cell = ws.cell(row=row, column=5, value=data.get("funnel", ""))
        funnel_cell.border = thin_border
        funnel_cell.alignment = center_align
        if funnel_type == "Awareness":
            funnel_cell.fill = awareness_fill
            funnel_cell.font = awareness_font
        elif funnel_type == "Consideration":
            funnel_cell.fill = consideration_fill
            funnel_cell.font = consideration_font
        else:
            funnel_cell.fill = conversion_fill
            funnel_cell.font = conversion_font

        # Column 6: Platform
        platform_cell = ws.cell(row=row, column=6, value=data.get("platform", ""))
        platform_cell.fill = platform_fill
        platform_cell.font = platform_font
        platform_cell.border = thin_border
        platform_cell.alignment = center_align

        # Column 7: Campaign / Ad Format
        c7 = ws.cell(row=row, column=7, value=data.get("campaign_adformat", ""))
        c7.border = thin_border
        c7.alignment = center_align
        c7.font = data_font

        # Budget (8-10)
        bp = data.get("budget_planned", 0)
        ba = data.get("budget_achieved", 0)
        bpct = data.get("budget_pct", 0)
        c8 = ws.cell(row=row, column=8, value=bp if bp else None)
        c8.number_format = "$#,##0"; c8.border = thin_border; c8.alignment = right_align; c8.font = data_font
        c9 = ws.cell(row=row, column=9, value=ba if ba else None)
        c9.number_format = "$#,##0.00"; c9.border = thin_border; c9.alignment = right_align; c9.font = data_font
        c10 = ws.cell(row=row, column=10, value=bpct if bpct else None)
        c10.number_format = "0%"; c10.border = thin_border; c10.alignment = center_align; c10.font = data_font

        # Impressions (11-13)
        ip = data.get("impressions_planned", 0)
        ia = data.get("impressions_achieved", 0)
        ipct = data.get("impressions_pct", 0)
        c11 = ws.cell(row=row, column=11, value=ip if ip else None)
        c11.number_format = "#,##0"; c11.border = thin_border; c11.alignment = right_align; c11.font = data_font
        c12 = ws.cell(row=row, column=12, value=ia if ia else None)
        c12.number_format = "#,##0"; c12.border = thin_border; c12.alignment = right_align; c12.font = data_font
        c13 = ws.cell(row=row, column=13, value=ipct if ipct else None)
        c13.number_format = "0%"; c13.border = thin_border; c13.alignment = center_align; c13.font = data_font

        # Metric 14-16 (Reach for Awareness, Clicks for Consideration)
        mp = data.get("metric_planned", 0)
        ma = data.get("metric_achieved", 0)
        mpct = data.get("metric_pct", 0)
        c14 = ws.cell(row=row, column=14, value=mp if mp else None)
        c14.number_format = "#,##0"; c14.border = thin_border; c14.alignment = right_align; c14.font = data_font
        c15 = ws.cell(row=row, column=15, value=ma if ma else None)
        c15.number_format = "#,##0"; c15.border = thin_border; c15.alignment = right_align; c15.font = data_font
        c16 = ws.cell(row=row, column=16, value=mpct if mpct else None)
        c16.number_format = "0%"; c16.border = thin_border; c16.alignment = center_align; c16.font = data_font

        # CPM/CPC (17-18)
        cpmp = data.get("cpm_planned", 0)
        cpma = data.get("cpm_achieved", 0)
        c17 = ws.cell(row=row, column=17, value=cpmp if cpmp else None)
        c17.number_format = "$#,##0.00"; c17.border = thin_border; c17.alignment = right_align; c17.font = data_font
        c18 = ws.cell(row=row, column=18, value=cpma if cpma else None)
        c18.number_format = "$#,##0.00"; c18.border = thin_border; c18.alignment = right_align; c18.font = data_font

        # Freq/CTR (19-20)
        fp = data.get("freq_planned", 0)
        fa = data.get("freq_achieved", 0)
        c19 = ws.cell(row=row, column=19, value=fp if fp else None)
        c19.number_format = "0.00"; c19.border = thin_border; c19.alignment = center_align; c19.font = data_font
        c20 = ws.cell(row=row, column=20, value=fa if fa else None)
        c20.number_format = "0.00"; c20.border = thin_border; c20.alignment = center_align; c20.font = data_font

        if funnel_type == "Awareness":
            c21 = ws.cell(row=row, column=21, value=data.get("views_achieved", 0) or None)
            c21.number_format = "#,##0"; c21.border = thin_border; c21.alignment = right_align; c21.font = data_font
            c22 = ws.cell(row=row, column=22, value=data.get("engagement_achieved", 0) or None)
            c22.number_format = "#,##0"; c22.border = thin_border; c22.alignment = right_align; c22.font = data_font
            c23 = ws.cell(row=row, column=23, value=data.get("vtr_achieved", 0) or None)
            c23.number_format = "0.00%"; c23.border = thin_border; c23.alignment = center_align; c23.font = data_font
            c24 = ws.cell(row=row, column=24, value=data.get("er_achieved", 0) or None)
            c24.number_format = "0.00%"; c24.border = thin_border; c24.alignment = center_align; c24.font = data_font
            c25 = ws.cell(row=row, column=25, value=data.get("sessions_achieved", 0) or None)
            c25.number_format = "#,##0"; c25.border = thin_border; c25.alignment = right_align; c25.font = data_font
            c26 = ws.cell(row=row, column=26, value=data.get("purchases_achieved", 0) or None)
            c26.number_format = "#,##0"; c26.border = thin_border; c26.alignment = right_align; c26.font = data_font
            c27 = ws.cell(row=row, column=27, value=data.get("purchases_value_achieved", 0) or None)
            c27.number_format = "$#,##0.00"; c27.border = thin_border; c27.alignment = right_align; c27.font = data_font
            c28 = ws.cell(row=row, column=28, value=data.get("purchase_roas_achieved", 0) or None)
            c28.number_format = "0.00"; c28.border = thin_border; c28.alignment = right_align; c28.font = data_font
        else:
            c21 = ws.cell(row=row, column=21, value=data.get("sessions_achieved", 0) or None)
            c21.number_format = "#,##0"; c21.border = thin_border; c21.alignment = right_align; c21.font = data_font
            c22 = ws.cell(row=row, column=22, value=data.get("purchases_achieved", 0) or None)
            c22.number_format = "#,##0"; c22.border = thin_border; c22.alignment = right_align; c22.font = data_font
            c23 = ws.cell(row=row, column=23, value=data.get("purchases_value_achieved", 0) or None)
            c23.number_format = "$#,##0.00"; c23.border = thin_border; c23.alignment = right_align; c23.font = data_font
            c24 = ws.cell(row=row, column=24, value=data.get("purchase_roas_achieved", 0) or None)
            c24.number_format = "0.00"; c24.border = thin_border; c24.alignment = right_align; c24.font = data_font
            for col in range(25, 29):
                c = ws.cell(row=row, column=col, value=None)
                c.border = thin_border
        return row + 1

    def add_market_section(start_row, market_name, market_data):
        # Calculate total rows needed
        total_rows = 0
        for funnel in ["Awareness", "Consideration", "Conversion"]:
            rows = market_data.get(funnel, [])
            if rows:
                total_rows += 1 + 2 + len(rows)  # KPI banner + 2 header rows + data rows
        if total_rows == 0:
            return start_row

        # Merge market name cell (column 1)
        market_start = start_row
        market_end = start_row + total_rows - 1
        ws.merge_cells(start_row=market_start, start_column=1, end_row=market_end, end_column=1)
        market_cell = ws.cell(row=market_start, column=1, value=market_name)
        market_cell.fill = market_fill
        market_cell.font = market_font
        market_cell.alignment = Alignment(horizontal="center", vertical="center")
        market_cell.border = thin_border

        current_row = start_row
        for funnel in ["Awareness", "Consideration", "Conversion"]:
            rows = market_data.get(funnel, [])
            if not rows:
                continue
            last_col = 28 if funnel == "Awareness" else 24
            current_row = add_kpi_banner(current_row, last_col)
            if funnel == "Awareness":
                current_row = apply_header_row(current_row, AWARENESS_HEADERS, AWARENESS_SUBHEADERS, last_col)
            else:
                current_row = apply_header_row(current_row, CONSIDERATION_HEADERS, CONSIDERATION_SUBHEADERS, last_col)
            for data in rows:
                current_row = write_data_row(current_row, data, funnel)
        return current_row

    # Write each market
    current_row = 1
    for market in ALL_MARKETS:
        if market in report_data_by_market and report_data_by_market[market]:
            current_row = add_market_section(current_row, market, report_data_by_market[market])

    # Write ALL section
    if all_data:
        current_row = add_market_section(current_row, "ALL", all_data)

    # Freeze panes at column E, row 3
    ws.freeze_panes = "H3"

    return wb


# =============================================================================
# STREAMLIT APP
# =============================================================================

st.header("1. Upload Files")
col_up1, col_up2 = st.columns(2)

with col_up1:
    uploaded_raw = st.file_uploader("Upload Raw Data CSV", type=["csv"], key="raw")

with col_up2:
    uploaded_media = st.file_uploader("Upload Media Plan Excel (optional)", type=["xlsx", "xls"], key="media")

media_plan_df = None
if uploaded_media is not None:
    try:
        xl = pd.ExcelFile(uploaded_media)
        media_sheet = None
        for sheet in xl.sheet_names:
            if "media plan" in sheet.lower():
                media_sheet = sheet
                break
        if media_sheet is None:
            media_sheet = xl.sheet_names[0]

        temp_df = pd.read_excel(uploaded_media, sheet_name=media_sheet, header=None)
        header_row = find_header_row(temp_df, "campaign name")
        media_plan_df = pd.read_excel(uploaded_media, sheet_name=media_sheet, header=header_row)
        media_plan_df.columns = [str(c).strip().lower() for c in media_plan_df.columns]
        st.success(f"Media plan loaded: {len(media_plan_df)} rows from sheet '{media_sheet}'")
    except Exception as e:
        st.error(f"Error reading media plan: {e}")
        media_plan_df = None

if uploaded_raw is not None:
    df = pd.read_csv(uploaded_raw)
    df["Week"] = pd.to_numeric(df["Week"], errors="coerce").astype("Int64")
    df["Placement"] = df["Campaign"].apply(extract_placement)

    available_brands = sorted(df["Brand"].dropna().unique())
    available_markets = sorted(df["Market"].dropna().unique())
    available_channels = sorted(df["Channel"].dropna().unique())
    available_funnels = sorted(df["Funnel Type"].dropna().unique())
    available_weeks = sorted(df["Week"].dropna().unique())

    st.success(f"Raw data loaded: {len(df):,} rows")

    # --- Filters ---
    st.header("2. Filter Options")
    col1, col2 = st.columns(2)
    with col1:
        campaign_search = st.text_input(
            "Campaign Name (or partial code)",
            placeholder="e.g. signature, asc, 2026...",
            help="Enter a keyword or code that appears in campaign names. Leave empty to include all."
        )
        if "week_selected" not in st.session_state:
            st.session_state.week_selected = [available_weeks[-1]] if available_weeks else []
        if st.button("Select All Weeks"):
            st.session_state.week_selected = available_weeks
            st.rerun()
        week_selected = st.multiselect("Week Number(s)", options=available_weeks, key="week_selected")
    with col2:
        st.write("")

    col3, col4 = st.columns(2)
    with col3:
        brand_options = available_brands if available_brands else ALL_BRANDS
        selected_brands = st.multiselect("Brand(s)", options=brand_options, default=brand_options[0] if brand_options else [])
        market_options = available_markets if available_markets else ALL_MARKETS
        selected_markets = st.multiselect("Market(s)", options=market_options, default=market_options)
    with col4:
        channel_options = available_channels if available_channels else ALL_CHANNELS
        selected_channels = st.multiselect("Channel(s)", options=channel_options, default=channel_options)
        funnel_options = available_funnels if available_funnels else ALL_FUNNELS
        selected_funnels = st.multiselect("Funnel Type(s)", options=funnel_options, default=funnel_options)

    combine_platforms = st.checkbox("Combine similar platforms", value=False,
        help="When checked, similar platforms are grouped together (e.g. TIKTOK REACH + TIKTOK PULSE -> TIKTOK)")

    # --- Apply Filters ---
    filtered_df = df.copy()
    if week_selected:
        filtered_df = filtered_df[filtered_df["Week"].isin(week_selected)]
    else:
        filtered_df = filtered_df.iloc[0:0]  # empty if no weeks selected
    if campaign_search.strip():
        search_term = campaign_search.strip().lower()
        filtered_df = filtered_df[filtered_df["Campaign"].str.lower().str.contains(search_term, na=False)]
    if selected_brands:
        filtered_df = filtered_df[filtered_df["Brand"].isin(selected_brands)]
    if selected_markets:
        filtered_df = filtered_df[filtered_df["Market"].isin(selected_markets)]
    if selected_channels:
        filtered_df = filtered_df[filtered_df["Channel"].isin(selected_channels)]
    if selected_funnels:
        filtered_df = filtered_df[filtered_df["Funnel Type"].isin(selected_funnels)]

    filtered_media = media_plan_df.copy() if media_plan_df is not None else None
    if filtered_media is not None and campaign_search.strip():
        search_term = campaign_search.strip().lower()
        filtered_media = filtered_media[
            filtered_media["campaign name"].astype(str).str.lower().str.contains(search_term, na=False)
        ]

    # --- Market Budget Split % ---
    st.header("3. Market Budget Split %")
    st.caption("Optional: enter budget split % per market. Used when no media plan is uploaded or to override planned budgets.")
    market_split = {}
    split_cols = st.columns(min(len(selected_markets), 4))
    for i, market in enumerate(selected_markets):
        with split_cols[i % 4]:
            market_split[market] = st.number_input(
                f"{market} %", min_value=0.0, max_value=100.0, value=0.0, step=1.0,
                key=f"split_{market}"
            ) / 100.0

    total_split = sum(market_split.values())
    if total_split > 0:
        st.info(f"Total split: {total_split*100:.1f}%")

    # --- Week Aggregation ---
    if len(week_selected) > 1:
        st.info(f"Aggregating data across {len(week_selected)} weeks: {week_selected}")
        # Aggregate by campaign (sum base metrics, recalculate derived)
        agg_cols = {
            "Year": "first", "Period": "first", "Week": lambda x: ", ".join(map(str, sorted(x.unique()))),
            "Brand": "first", "Market": "first", "Channel": "first", "Funnel Type": "first",
            "Campaign": "first", "Placement": "first",
            "Cost $": "sum", "Clicks": "sum", "Impressions": "sum",
            "Reach": "max", "Full Video Views": "sum", "Sessions": "sum",
            "Orders": "sum", "Revenue $": "sum", "Off. Orders": "sum", "Off. Revenue $": "sum",
            "App Installs": "sum",
        }
        filtered_df = filtered_df.groupby("Campaign", as_index=False).agg(agg_cols)
        # Recalculate derived metrics
        filtered_df["CPC"] = filtered_df["Cost $"] / filtered_df["Clicks"].replace(0, pd.NA)
        filtered_df["CPM"] = (filtered_df["Cost $"] / filtered_df["Impressions"].replace(0, pd.NA)) * 1000
        filtered_df["CTR"] = filtered_df["Clicks"] / filtered_df["Impressions"].replace(0, pd.NA)
        filtered_df["Frequency"] = filtered_df["Impressions"] / filtered_df["Reach"].replace(0, pd.NA)
        filtered_df["VTR"] = filtered_df["Full Video Views"] / filtered_df["Impressions"].replace(0, pd.NA)
        filtered_df["Click To Session Rate"] = filtered_df["Sessions"] / filtered_df["Clicks"].replace(0, pd.NA)
        filtered_df["CvR"] = filtered_df["Orders"] / filtered_df["Sessions"].replace(0, pd.NA)

    st.header("4. Preview Filtered Data")
    weeks_label = ", ".join(map(str, week_selected)) if week_selected else "None"
    st.write(f"**Weeks:** {weeks_label}  |  **Raw data rows:** {len(filtered_df):,}")
    if filtered_media is not None:
        st.write(f"**Media plan rows:** {len(filtered_media):,}")

    if len(filtered_df) > 0:
        st.dataframe(filtered_df, use_container_width=True)

        # --- Summary Views ---
        st.header("4. Summary Views")
        tab1, tab2, tab3 = st.tabs(["By Market", "By Channel", "By Funnel"])
        for tab, group_col in [(tab1, "Market"), (tab2, "Channel"), (tab3, "Funnel Type")]:
            with tab:
                # Reach is unique users: use max at channel level (don't double-count),
                # but sum at market/funnel level since those are different audiences.
                reach_agg = "max" if group_col == "Channel" else "sum"
                summary = filtered_df.groupby(group_col).agg({
                    "Cost $": "sum", "Impressions": "sum", "Reach": reach_agg,
                    "Clicks": "sum", "Sessions": "sum", "Orders": "sum",
                    "Revenue $": "sum", "Off. Orders": "sum", "Off. Revenue $": "sum"
                }).reset_index()
                summary["Total Orders"] = summary["Orders"] + summary["Off. Orders"]
                summary["Total Revenue"] = summary["Revenue $"] + summary["Off. Revenue $"]
                summary["ROAS"] = summary["Total Revenue"] / summary["Cost $"].replace(0, pd.NA)
                summary["CvR"] = summary["Total Orders"] / summary["Sessions"].replace(0, pd.NA)
                st.dataframe(summary.round(2), use_container_width=True)

        # --- Generate Report ---
        st.header("5. Generate Report")
        if st.button("📥 Generate Excel Report", type="primary"):
            with st.spinner("Generating report..."):
                # Compute derived metrics
                filtered_df = filtered_df.copy()
                filtered_df["Total Revenue"] = filtered_df["Revenue $"] + filtered_df["Off. Revenue $"]
                filtered_df["Total Orders"] = filtered_df["Orders"] + filtered_df["Off. Orders"]
                filtered_df["ROAS"] = filtered_df["Total Revenue"] / filtered_df["Cost $"].replace(0, pd.NA)

                # Build media plan lookup
                media_lookup = {}
                if filtered_media is not None:
                    for _, row in filtered_media.iterrows():
                        norm = normalize_campaign(row["campaign name"])
                        media_lookup[norm] = row

                # --- Market Budget Split Calculation ---
                # Pre-compute market-level achieved budgets for split allocation
                total_achieved_budget = filtered_df["Cost $"].sum()
                market_achieved_budget = filtered_df.groupby("Market")["Cost $"].sum().to_dict()
                market_campaign_count = filtered_df.groupby("Market").size().to_dict()

                # Build report data
                report_rows = []
                for _, row in filtered_df.iterrows():
                    norm = normalize_campaign(row["Campaign"])
                    planned = media_lookup.get(norm)
                    if planned is None:
                        for mk, mv in media_lookup.items():
                            if norm in mk or mk in norm:
                                planned = mv
                                break

                    funnel = row["Funnel Type"]
                    budget_planned = 0
                    impressions_planned = 0
                    reach_planned = 0
                    market = row["Market"]
                    channel = row["Channel"]
                    optimization = ""

                    # Media plan fields
                    region = ""
                    audience = ""
                    placement = ""

                    if planned is not None:
                        budget_planned = safe_numeric(planned.get("budget  usd", 0))
                        impressions_planned = safe_numeric(planned.get("impressions", 0))
                        reach_planned = safe_numeric(planned.get("reach", 0))
                        market = str(planned.get("market", row["Market"])).strip().upper()
                        channel = planned.get("channel", row["Channel"])
                        optimization = planned.get("optimization", "")
                        region = str(planned.get("region", "")).strip()
                        audience = str(planned.get("audience", "")).strip()
                        placement = str(planned.get("placement", "")).strip()

                    # If no media plan budget but user entered market split %, calculate planned budget
                    budget_achieved = row["Cost $"]
                    if budget_planned == 0 and total_achieved_budget > 0:
                        split_pct = market_split.get(market, 0)
                        if split_pct > 0:
                            market_budget = total_achieved_budget * split_pct
                            market_ach = market_achieved_budget.get(market, 0)
                            if market_ach > 0:
                                budget_planned = market_budget * (budget_achieved / market_ach)

                    impressions_achieved = row["Impressions"]
                    reach_achieved = row["Reach"]
                    clicks_achieved = row["Clicks"]
                    sessions_achieved = row["Sessions"]
                    orders_achieved = row["Orders"]
                    revenue_achieved = row["Revenue $"]
                    roas = row["ROAS"] if pd.notna(row["ROAS"]) else 0

                    budget_pct = (budget_achieved / budget_planned) if budget_planned else 0
                    impressions_pct = (impressions_achieved / impressions_planned) if impressions_planned else 0
                    reach_pct = (reach_achieved / reach_planned) if reach_planned else 0
                    clicks_pct = 0

                    cpm_planned = (budget_planned / impressions_planned * 1000) if impressions_planned else 0
                    cpm_achieved = (budget_achieved / impressions_achieved * 1000) if impressions_achieved else 0
                    cpc_achieved = (budget_achieved / clicks_achieved) if clicks_achieved else 0
                    ctr_achieved = row["CTR"] if pd.notna(row["CTR"]) else 0
                    # Frequency = Impressions / Reach (calculated, not taken from raw data)
                    freq_planned = (impressions_planned / reach_planned) if reach_planned else 0
                    freq_achieved = (impressions_achieved / reach_achieved) if reach_achieved else 0
                    views_achieved = row["Full Video Views"] if pd.notna(row["Full Video Views"]) else 0
                    vtr_achieved = row["VTR"] if pd.notna(row["VTR"]) else 0
                    engagement_achieved = row["Sessions"] if pd.notna(row["Sessions"]) else 0

                    # ER = Engagement / Impressions (proxy using Sessions)
                    er_achieved = (engagement_achieved / impressions_achieved) if impressions_achieved else 0

                    # Fallback: extract placement from campaign name if not in media plan
                    if not placement:
                        placement = extract_placement(row["Campaign"])

                    platform = format_platform(channel, optimization, funnel, row["Campaign"], combine=combine_platforms)
                    campaign_adformat = str(optimization).strip() if optimization else ""

                    if funnel == "Awareness":
                        report_rows.append({
                            "market": market,
                            "funnel": funnel,
                            "platform": platform,
                            "campaign_adformat": campaign_adformat,
                            "budget_planned": budget_planned,
                            "budget_achieved": budget_achieved,
                            "budget_pct": budget_pct,
                            "impressions_planned": impressions_planned,
                            "impressions_achieved": impressions_achieved,
                            "impressions_pct": impressions_pct,
                            "metric_planned": reach_planned,
                            "metric_achieved": reach_achieved,
                            "metric_pct": reach_pct,
                            "cpm_planned": cpm_planned,
                            "cpm_achieved": cpm_achieved,
                            "freq_planned": freq_planned,
                            "freq_achieved": freq_achieved,
                            "views_achieved": views_achieved,
                            "engagement_achieved": engagement_achieved,
                            "vtr_achieved": vtr_achieved,
                            "er_achieved": er_achieved,
                            "sessions_achieved": sessions_achieved,
                            "purchases_achieved": orders_achieved,
                            "purchases_value_achieved": revenue_achieved,
                            "purchase_roas_achieved": roas,
                            "placement": placement,
                            "region": region,
                            "audience": audience,
                        })
                    else:
                        # Consideration or Conversion
                        report_rows.append({
                            "market": market,
                            "funnel": funnel,
                            "platform": platform,
                            "campaign_adformat": campaign_adformat,
                            "budget_planned": budget_planned,
                            "budget_achieved": budget_achieved,
                            "budget_pct": budget_pct,
                            "impressions_planned": impressions_planned,
                            "impressions_achieved": impressions_achieved,
                            "impressions_pct": impressions_pct,
                            "metric_planned": 0,
                            "metric_achieved": clicks_achieved,
                            "metric_pct": 0,
                            "cpm_planned": 0,
                            "cpm_achieved": cpc_achieved,
                            "freq_planned": 0,
                            "freq_achieved": ctr_achieved,
                            "views_achieved": 0,
                            "engagement_achieved": 0,
                            "vtr_achieved": 0,
                            "er_achieved": 0,
                            "sessions_achieved": sessions_achieved,
                            "purchases_achieved": orders_achieved,
                            "purchases_value_achieved": revenue_achieved,
                            "purchase_roas_achieved": roas,
                            "placement": placement,
                            "region": region,
                            "audience": audience,
                        })

                # Group by market and funnel
                report_data_by_market = {}
                for r in report_rows:
                    m = r["market"]
                    f = r["funnel"]
                    if m not in report_data_by_market:
                        report_data_by_market[m] = {}
                    if f not in report_data_by_market[m]:
                        report_data_by_market[m][f] = []
                    report_data_by_market[m][f].append(r)

                # Build ALL aggregation
                all_data = {}
                # Track max reach per market per platform so we sum across markets
                # but don't double-count campaigns within the same market+channel.
                awareness_market_reach = {}  # key -> {market: {"planned": x, "achieved": y}}
                for r in report_rows:
                    f = r["funnel"]
                    key = (f, r["platform"], r["campaign_adformat"])
                    if f not in all_data:
                        all_data[f] = {}
                    if key not in all_data[f]:
                        all_data[f][key] = {
                            "market": "ALL",
                            "funnel": f,
                            "platform": r["platform"],
                            "campaign_adformat": r["campaign_adformat"],
                            "region": "",
                            "audience": "",
                            "placement": "",
                            "budget_planned": 0,
                            "budget_achieved": 0,
                            "impressions_planned": 0,
                            "impressions_achieved": 0,
                            "metric_planned": 0,
                            "metric_achieved": 0,
                            "cpm_planned": 0,
                            "cpm_achieved": 0,
                            "freq_planned": 0,
                            "freq_achieved": 0,
                            "views_achieved": 0,
                            "engagement_achieved": 0,
                            "vtr_achieved": 0,
                            "er_achieved": 0,
                            "sessions_achieved": 0,
                            "purchases_achieved": 0,
                            "purchases_value_achieved": 0,
                            "purchase_roas_achieved": 0,
                        }
                        if f == "Awareness":
                            awareness_market_reach[key] = {}
                    entry = all_data[f][key]
                    if not entry["region"] and r.get("region"):
                        entry["region"] = r["region"]
                    if not entry["audience"] and r.get("audience"):
                        entry["audience"] = r["audience"]
                    if not entry["placement"] and r.get("placement"):
                        entry["placement"] = r["placement"]
                    entry["budget_planned"] += r["budget_planned"]
                    entry["budget_achieved"] += r["budget_achieved"]
                    entry["impressions_planned"] += r["impressions_planned"]
                    entry["impressions_achieved"] += r["impressions_achieved"]
                    entry["sessions_achieved"] += r["sessions_achieved"]
                    entry["purchases_achieved"] += r["purchases_achieved"]
                    entry["purchases_value_achieved"] += r["purchases_value_achieved"]
                    if f == "Awareness":
                        market = r["market"]
                        if market not in awareness_market_reach[key]:
                            awareness_market_reach[key][market] = {"planned": 0, "achieved": 0}
                        awareness_market_reach[key][market]["planned"] = max(awareness_market_reach[key][market]["planned"], r["metric_planned"])
                        awareness_market_reach[key][market]["achieved"] = max(awareness_market_reach[key][market]["achieved"], r["metric_achieved"])
                        entry["metric_planned"] = sum(v["planned"] for v in awareness_market_reach[key].values())
                        entry["metric_achieved"] = sum(v["achieved"] for v in awareness_market_reach[key].values())
                        entry["views_achieved"] += r["views_achieved"]
                        entry["engagement_achieved"] += r["engagement_achieved"]
                    else:
                        entry["metric_planned"] += r["metric_planned"]
                        entry["metric_achieved"] += r["metric_achieved"]

                # Compute derived fields for ALL
                all_report_data = {}
                for f, entries in all_data.items():
                    all_report_data[f] = []
                    for entry in entries.values():
                        entry["budget_pct"] = (entry["budget_achieved"] / entry["budget_planned"]) if entry["budget_planned"] else 0
                        entry["impressions_pct"] = (entry["impressions_achieved"] / entry["impressions_planned"]) if entry["impressions_planned"] else 0
                        entry["metric_pct"] = (entry["metric_achieved"] / entry["metric_planned"]) if entry["metric_planned"] else 0
                        entry["cpm_planned"] = (entry["budget_planned"] / entry["impressions_planned"] * 1000) if entry["impressions_planned"] else 0
                        entry["cpm_achieved"] = (entry["budget_achieved"] / entry["impressions_achieved"] * 1000) if entry["impressions_achieved"] else 0
                        entry["purchase_roas_achieved"] = (entry["purchases_value_achieved"] / entry["budget_achieved"]) if entry["budget_achieved"] else 0
                        if f == "Awareness":
                            entry["vtr_achieved"] = (entry["views_achieved"] / entry["impressions_achieved"]) if entry["impressions_achieved"] else 0
                            entry["er_achieved"] = (entry["engagement_achieved"] / entry["impressions_achieved"]) if entry["impressions_achieved"] else 0
                            # Frequency = Impressions / Reach for ALL aggregation
                            entry["freq_planned"] = (entry["impressions_planned"] / entry["metric_planned"]) if entry["metric_planned"] else 0
                            entry["freq_achieved"] = (entry["impressions_achieved"] / entry["metric_achieved"]) if entry["metric_achieved"] else 0
                        all_report_data[f].append(entry)

                # Build formatted REPORT workbook
                wb = build_formatted_report(report_data_by_market, all_report_data)

                # Add data sheets using pandas
                raw_data = pd.DataFrame({
                    "Campaign name": filtered_df["Campaign"],
                    "Placement": filtered_df["Placement"],
                    "Cost": filtered_df["Cost $"],
                    "Impressions": filtered_df["Impressions"],
                    "Reach": filtered_df["Reach"],
                    "Clicks (all)": filtered_df["Clicks"],
                    "2-second video views": filtered_df["Full Video Views"],
                    "Frequency": filtered_df["Frequency"],
                    "ENGAGEMENT": filtered_df["Sessions"],
                    "LANDING PAGE VIEWS": filtered_df["Sessions"],
                    "Purchases": filtered_df["Orders"],
                    "Purchases Value": filtered_df["Revenue $"],
                    "Purchase Roas": filtered_df["ROAS"].fillna(0),
                })

                detail_data = filtered_df.copy()

                summary = filtered_df.groupby(["Market", "Channel", "Funnel Type"]).agg({
                    "Cost $": "sum", "Impressions": "sum", "Reach": "max",
                    "Clicks": "sum", "Sessions": "sum", "Orders": "sum",
                    "Revenue $": "sum", "Off. Orders": "sum", "Off. Revenue $": "sum"
                }).reset_index()
                summary["Total Orders"] = summary["Orders"] + summary["Off. Orders"]
                summary["Total Revenue"] = summary["Revenue $"] + summary["Off. Revenue $"]
                summary["ROAS"] = summary["Total Revenue"] / summary["Cost $"].replace(0, pd.NA)
                summary["CvR"] = summary["Total Orders"] / summary["Sessions"].replace(0, pd.NA)

                # Helper: replace pd.NA with None for openpyxl compatibility
                def clean_df(df):
                    return df.astype(object).where(pd.notna(df), None)

                # Write MARKET SPLIT sheet
                ws_split = wb.create_sheet("MARKET SPLIT")
                split_data = []
                for m in selected_markets:
                    split_data.append({
                        "Market": m,
                        "Budget Split %": market_split.get(m, 0) * 100,
                        "Achieved Budget": market_achieved_budget.get(m, 0),
                        "Achieved % of Total": (market_achieved_budget.get(m, 0) / total_achieved_budget * 100) if total_achieved_budget else 0,
                    })
                split_df = pd.DataFrame(split_data)
                for r in dataframe_to_rows(clean_df(split_df), index=False, header=True):
                    ws_split.append(r)

                # Write data sheets to the same workbook
                ws_raw = wb.create_sheet("RAW")
                for r in dataframe_to_rows(clean_df(raw_data), index=False, header=True):
                    ws_raw.append(r)

                ws_detail = wb.create_sheet("DETAIL")
                for r in dataframe_to_rows(clean_df(detail_data), index=False, header=True):
                    ws_detail.append(r)

                ws_summary = wb.create_sheet("SUMMARY")
                for r in dataframe_to_rows(clean_df(summary), index=False, header=True):
                    ws_summary.append(r)

                if filtered_media is not None:
                    ws_media = wb.create_sheet("MEDIA PLAN")
                    media_out = filtered_media.copy()
                    media_out = media_out.drop(columns=["_norm_name"], errors="ignore")
                    for r in dataframe_to_rows(clean_df(media_out), index=False, header=True):
                        ws_media.append(r)

                for funnel in filtered_df["Funnel Type"].unique():
                    funnel_df = filtered_df[filtered_df["Funnel Type"] == funnel]
                    pivot = funnel_df.groupby(["Market", "Channel"]).agg({
                        "Cost $": "sum", "Impressions": "sum", "Reach": "max",
                        "Clicks": "sum", "Sessions": "sum", "Orders": "sum",
                        "Revenue $": "sum", "Off. Orders": "sum", "Off. Revenue $": "sum"
                    }).reset_index()
                    pivot["Total Orders"] = pivot["Orders"] + pivot["Off. Orders"]
                    pivot["Total Revenue"] = pivot["Revenue $"] + pivot["Off. Revenue $"]
                    pivot["ROAS"] = pivot["Total Revenue"] / pivot["Cost $"].replace(0, pd.NA)
                    pivot["CvR"] = pivot["Total Orders"] / pivot["Sessions"].replace(0, pd.NA)
                    ws_funnel = wb.create_sheet(funnel.upper()[:10])
                    for r in dataframe_to_rows(clean_df(pivot), index=False, header=True):
                        ws_funnel.append(r)

                # Move REPORT to front
                sheets = wb._sheets
                report_sheet = wb["REPORT"]
                sheets.remove(report_sheet)
                sheets.insert(0, report_sheet)

                output = BytesIO()
                wb.save(output)
                output.seek(0)

                brands_str = "_".join(selected_brands) if selected_brands else "All"
                markets_str = "_".join(selected_markets) if selected_markets else "All"
                weeks_str = "_".join(map(str, week_selected)) if week_selected else "None"
                filename = f"Inflight_Report_W{weeks_str}_{brands_str}_{markets_str}.xlsx"

                st.download_button(
                    label="⬇️ Download Excel Report",
                    data=output,
                    file_name=filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                st.success("Report generated successfully!")
    else:
        st.warning("No data matches the selected filters. Please adjust your criteria.")
else:
    st.info("Please upload a Raw Data CSV file to begin.")
